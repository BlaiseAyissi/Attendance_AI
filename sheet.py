import openpyxl.workbook
import xlwt
import xlrd
from xlutils.copy import copy
import openpyxl
import os
import datetime
import xlsxwriter

st_name = 'Aashish'
def mark_present(st_name):

	names = os.listdir('output/')
	print(names)

	sub = 'SAMPLE'
	filepath = 'attendance/' + sub + '.xlsx'
	
	if not os.path.exists(filepath):
		count = 2
		workbook = openpyxl.Workbook()
		print("Creating Spreadsheet with Title: " + sub)
		sheet = workbook.active 
		for i in names:
			sheet.cell(row=count, column=1, value=i)
			count += 1

		workbook.save(filepath) 

	wb = openpyxl.load_workbook(filepath)
	#wb = copy(rb)
	sheet = wb.active
	sheet.cell(row=1,column=1,value=str(datetime.datetime.now()))


	count = 2
	for i in names:
		if i in st_name:
			sheet.cell(row=count, column=2, value='P')
		else:
			sheet.cell(row=count, column=2, value='A')

		sheet.cell(row=count, column=1, value=i)
		count += 1

	wb.save(filepath)


mark_present(st_name)
